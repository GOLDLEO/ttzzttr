import openpyxl
from openpyxl.workbook import Workbook
from pathlib import Path
import re
import requests
import time 

def fix_time():
	return time.perf_counter()

def fix_end_time(start_time):
	fix_now = time.perf_counter()
	result = f'{fix_now - start_time:0.4f}'
	return f"Прошло {result} секунд | {round(float(result) / 60, 2) } минут."

t_start = fix_time()



def open_and_read_xlsx(file_name='444.xlsx'):	
	xlsx_file = Path(file_name)
	wb_obj = openpyxl.load_workbook(xlsx_file)
	sheet = wb_obj.active

	col_names = []
	for column in sheet.iter_cols(1, sheet.max_column):
	    col_names.append(column[0].value)

#toc = time.perf_counter()

#c_time = f'{toc - t_start:0.4f}'
#print(f"Прошло {c_time} секунд | {float(c_time) / 60 } минут.")


def send_request(ip):
	url = f'http://ip-api.com/json/{ip}'
	while True:
		try:
			r = requests.get(url)
			data = r.json()
			print(data)
			if data['query'] == 'IP голосування':
				break
			if data['status'] == 'fail':
				print('Wait for sleep...')
				time.sleep(15)
			elif data['status'] == 'success':
				return data
		except:
			time.sleep(5)
			continue


def write_data_from_xlsx():
	field_country_code = []
	c = 40
	for row in sheet.iter_rows():
		c = c - 1 
		if c == 40:

			c = 40
			continue
		data = send_request(ip = row[1].value)
		#print(data)
		try:
			print(data['countryCode'])
			print(2)
			field_country_code.append(data['countryCode'])
			print(3)
		except Exception as err:
			print('===>', err)


def write_data_to_file(file_name='countryCode.txt'):	
	i = 0
	file = open(file_name, 'a')
	for field in field_country_code:
		field = str(field) + '\n' 
		file.write(field)

	file.close()

def read_file(file_name):
	l_file_fields = []
	with open(file_name, 'r') as f:
		for field in f.readlines():
			l_file_fields.append(field.replace('\n', ''))
	return l_file_fields


def write_data_to_xlsx(file_name, field_title='N1', l_fields=None):
	wb = openpyxl.load_workbook(file_name)
	sheet = wb.active
	create_new_field = sheet[field_title]
	create_new_field.value = 'Страна'
	if type(l_fields) == list:
		print(True)
		for idx, field in enumerate(l_fields, start=2):
			field_name = f'{field_title[0]}{int(idx)}'
			add_new_field = sheet[field_name]
			add_new_field.value = field
	try:
		wb.save(file_name)
		return 'success save'
	except Exception as err:
		return err



print(write_data_to_xlsx('444__test.xlsx', l_fields=read_file('countryCode.txt')))
print(fix_end_time(t_start))

#toc = time.perf_counter()

#time = f'{toc - tic_first:0.4f}'
#print(f"Всего на обработку ушло: {time} секунд | {float(time) / 60 } минут.")






